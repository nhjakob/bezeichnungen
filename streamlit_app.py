#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Streamlit Slot Selector (3 Raeder) mit:
- Up/Down-Buttons pro Rad (zentriert, volle Breite)
- "Spin" mit weichem Auslaufen (Slot-Feeling)
- Ergebniszeile als Berufsbezeichnung (Level [+], Leistung + Rolle aus Disziplin)
- Regeln für "nicht verfügbar": aus Excel (Liste: Status=nicht verfügbar / Matrix: rot markiert)
- 'Mid-Level' wird im Ergebnis unterdrueckt
- THEME an neues handeln angelehnt

NEU:
- Disziplin "Research" und "Projektmanagement" entfernt
- Sidebar-Upload zum Einlesen eurer Excel-Regeln
"""

import streamlit as st
import time
import random
from io import BytesIO

# Excel-Parsing
from openpyxl import load_workbook
# -------------------------
# Seiten-Setup
# -------------------------
st.set_page_config(
    page_title="Positionsbezeichnungen bei neues handeln",
    page_icon="🎰",
    layout="wide",
)

# THEME
THEME = {
    "bg": "#FFFFFF",
    "surface": "#F5F7FA",
    "text": "#0F172A",
    "muted": "#374151",
    "border": "#D1D5DB",
    "accent": "#00A886",
    "accent_surface": "#E6FFFA",
    "result_bg": "#0F766E",
}

# CSS
st.markdown(
    f"""
    <style>
    .result-bar {{
        background: {THEME["result_bg"]};
        color: #fff;
        padding: 14px 18px;
        border-radius: 12px;
        font-weight: 800;
        font-size: 1.12rem;
        margin: 16px auto 10px auto;
        letter-spacing: .1px;
        text-align: center;
        max-width: 900px;
    }}
    .value-box {{
        background: #fff;
        border: 1px solid {THEME["border"]};
        border-radius: 12px;
        padding: 14px 16px;
        text-align: center;
        font-weight: 700;
        color: {THEME["text"]};
        box-shadow: 0 0 0 4px {THEME["accent_surface"]};
    }}
    .title {{
        font-weight: 800;
        color: {THEME["text"]};
        margin-bottom: 8px;
        font-size: 1.05rem;
        letter-spacing: .2px;
        text-align: center;
    }}
    /* Primäre Buttons (z. B. Spin) */
    .stButton > button[kind="primary"] {{
        background-color: {THEME["accent"]} !important;
        color: #ffffff !important;
        border: 0;
        border-radius: 12px;
        padding: 10px 14px;
        font-weight: 800;
        letter-spacing: .2px;
    }}
    /* Sekundäre Buttons (Up/Down, Reset) */
    .btn-secondary > button {{
        background-color: #F3F4F6 !important;
        color: {THEME["text"]} !important;
        border: 1px solid {THEME["border"]} !important;
        border-radius: 10px !important;
        padding: 8px 12px !important;
        font-weight: 700 !important;
        width: 100% !important;
        display: block !important;
        margin: 0 auto !important;
    }}
    /* Action-Buttons (Spin/Reset) zentriert und breit */
    .action-btn > button {{
        width: 100% !important;
        border-radius: 12px !important;
        padding: 10px 14px !important;
        font-weight: 800 !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

# -------------------------
# Kategorien & Werte
# (Research/Projektmanagement entfernt)
# -------------------------
CATEGORIES = [
    ("Level", ["Junior", "Mid-Level", "Senior"]),
    ("Leistung", [
        "Event", "Content", "Video", "Audio", "Social Media", "Dialog",
        "Kommunikation", "Wissenschaft", "Print", "Kampagne", "Web",
        "Presse/PR", "UX/UI",
    ]),
    ("Disziplin", [
        "Beratung", "Design", "Konzept", "Text", "Analyse", "Redaktion",
        # "Research", "Projektmanagement" -> entfernt
    ]),
]

# Disziplin → Rollen
ROLE_TITLE = {
    "Beratung": "Berater:in",
    "Design": "Designer:in",
    "Konzept": "Konzeptioner:in",
    "Text": "Texter:in",
    "Analyse": "Analyst:in",
    "Redaktion": "Redakteur:in",
}
ROLE_STEM = {
    "Beratung": "berater",
    "Design": "designer",
    "Konzept": "konzeptioner",
    "Text": "texter",
    "Analyse": "analyst",
    "Redaktion": "redakteur",
}
DEFAULT_ROLE_TITLE = "Spezialist:in"
DEFAULT_ROLE_STEM = "spezialist"

# -------------------------
# Regeln „nicht verfügbar“ (werden via Excel geladen)
# Format: {"level": "*", "leistung": "<>", "disziplin": "<>"}
# -------------------------
if "UNAVAILABLE_RULES" not in st.session_state:
    st.session_state.UNAVAILABLE_RULES = []  # wird per Upload befüllt
def _match_field(value: str, pattern: str | None) -> bool:
    return pattern in (None, "*") or value == pattern

def is_unavailable(level: str, leistung: str, disziplin: str) -> bool:
    for rule in st.session_state.UNAVAILABLE_RULES:
        if (_match_field(level, rule.get("level")) and
            _match_field(leistung, rule.get("leistung")) and
            _match_field(disziplin, rule.get("disziplin"))):
            return True
    return False

# -------------------------
# Hilfsfunktionen
# -------------------------
def clamp_mod(i: int, n: int) -> int:
    return 0 if n == 0 else (i % n)

def hyphenize(term: str) -> str:
    """Mehrwort-Leistung mit Bindestrich, z. B. 'Social Media' -> 'Social-Media'."""
    return "-".join(term.split())

def compose_title(level: str, leistung: str, disziplin: str) -> str:
    """
    Ergebnislogik:
      - 'Mid-Level' wird unterdrueckt
      - Wissenschaft + Text -> Wissenschaftstexter:in
      - Wissenschaft + <Disziplin> -> Wissenschafts<rollenstamm>:in
      - Kommunikation + <Disziplin> -> Kommunikations<rollenstamm>:in
      - Kampagne: Design/Beratung speziell, sonst Kampagnen-<Rolle>
      - Web: Design/Konzept speziell, sonst Web-<Rolle>
      - Presse/PR: Beratung/Text speziell (PR-/Presse-), sonst PR-<Rolle>
      - UX/UI: Design speziell, sonst UX/UI-<Rolle>
      - sonst: <Leistung-mit-Bindestrich>-<Rolle>
    """
    role_title = ROLE_TITLE.get(disziplin, DEFAULT_ROLE_TITLE)
    role_stem  = ROLE_STEM.get(disziplin, DEFAULT_ROLE_STEM)

    level_prefix = "" if level == "Mid-Level" else (level + " ")

    if leistung == "Wissenschaft":
        if disziplin == "Text":
            return f"{level_prefix}Wissenschaftstexter:in"
        return f"{level_prefix}Wissenschafts{role_stem}:in"

    if leistung == "Kommunikation":
        return f"{level_prefix}Kommunikations{role_stem}:in"

    if leistung == "Kampagne":
        if disziplin == "Design":
            return f"{level_prefix}Kampagnen-Designer:in"
        if disziplin == "Beratung":
            return f"{level_prefix}Kampagnen-Berater:in"
        return f"{level_prefix}Kampagnen-{role_title}"

    if leistung == "Web":
        if disziplin == "Design":
            return f"{level_prefix}Web-Designer:in"
        if disziplin == "Konzept":
            return f"{level_prefix}Web-Konzeptioner:in"
        return f"{level_prefix}Web-{role_title}"

    if leistung == "Presse/PR":
        if disziplin == "Beratung":
            return f"{level_prefix}PR-Berater:in"
        if disziplin == "Text":
            return f"{level_prefix}Presse-Texter:in"
        return f"{level_prefix}PR-{role_title}"

    if leistung == "UX/UI":
        if disziplin == "Design":
            return f"{level_prefix}UX/UI-Designer:in"
        return f"{level_prefix}UX/UI-{role_title}"

    leistung_h = hyphenize(leistung)
    return f"{level_prefix}{leistung_h}-{role_title}"

# -------------------------
# Excel-Import-Funktionen
# -------------------------
def parse_rules_from_excel(file_bytes: bytes) -> list[dict]:
    """
    Liest verbotene Kombinationen aus:
      - Sheet 'Liste': Status == 'nicht verfügbar'
      - Sheet 'Matrix': Zellen mit 'solid' Fill (z. B. rot markiert)
    Liefert Regeln mit level='*'.
    """
    wb = load_workbook(filename=BytesIO(file_bytes))
    rules = set()

    # --- Liste ---
    if 'Liste' in wb.sheetnames:
        ws = wb['Liste']
        # Header-Mapping
        header_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_map[str(val).strip().lower()] = c
        c_l = header_map.get('leistung')
        c_d = header_map.get('disziplin')
        c_s = header_map.get('status')

        if c_l and c_d and c_s:
            for r in range(2, ws.max_row + 1):
                l = ws.cell(row=r, column=c_l).value
                d = ws.cell(row=r, column=c_d).value
                s = ws.cell(row=r, column=c_s).value
                if l and d and s and str(s).strip().lower() == 'nicht verfügbar':
                    rules.add((str(l).strip(), str(d).strip()))

    # --- Matrix ---
    if 'Matrix' in wb.sheetnames:
        ws = wb['Matrix']
        # Disziplin-Header (Zeile 1, ab Spalte 2)
        disziplin = []
        col = 2
        while True:
            v = ws.cell(row=1, column=col).value
            if not v:
                break
            disziplin.append(str(v).strip())
            col += 1
        # Leistung-Header (Spalte 1, ab Zeile 2)
        leistung = []
        row = 2
        while True:
            v = ws.cell(row=row, column=1).value
            if not v:
                break
            leistung.append(str(v).strip())
            row += 1
        # Zellen prüfen
        for i, l in enumerate(leistung, start=2):
            for j, d in enumerate(disziplin, start=2):
                cell = ws.cell(row=i, column=j)
                fill = cell.fill
                marked = False
                if fill and getattr(fill, 'fill_type', None) == 'solid':
                    fg = getattr(fill.fgColor, 'rgb', None)
                    if fg:
                        rgb = fg.upper()
                        # Nicht-weiß behandeln wir als Markierung
                        if rgb not in ('FFFFFFFF', 'FF000000', '00000000'):
                            marked = True
                    else:
                        # indexed/Theme-Farben -> als markiert behandeln
                        marked = True
                # Optionales "X" als Marker
                if not marked and isinstance(cell.value, str) and cell.value.strip().lower() == 'x':
                    marked = True
                if marked:
                    rules.add((l, d))

    # In Regel-Form bringen
    return [{"level": "*", "leistung": l, "disziplin": d} for (l, d) in sorted(rules)]

# -------------------------
# State
# -------------------------
if "indices" not in st.session_state:
    st.session_state.indices = [0 for _ in CATEGORIES]

# -------------------------
# UI-Bausteine
# -------------------------
def value_html(text: str) -> str:
    return f'<div class="value-box">{text}</div>'

def render_wheels() -> list:
    """Rendert die 3 Räder nebeneinander und gibt Value-Placeholders zurück."""
    cols = st.columns(len(CATEGORIES), gap="large")
    value_placeholders = []
    for i, (title, options) in enumerate(CATEGORIES):
        with cols[i]:
            st.markdown(f'<div class="title">{title}</div>', unsafe_allow_html=True)

            # Up (zentriert, volle Breite)
            st.markdown('<div class="btn-secondary">', unsafe_allow_html=True)
            up_clicked = st.button("↑", key=f"up_{i}", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

            # Wert
            value_ph = st.empty()
            idx = clamp_mod(st.session_state.indices[i], len(options))
            st.session_state.indices[i] = idx
            value_ph.markdown(value_html(options[idx]), unsafe_allow_html=True)

            # Down (zentriert, volle Breite)
            st.markdown('<div class="btn-secondary">', unsafe_allow_html=True)
            down_clicked = st.button("↓", key=f"down_{i}", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

            # Interaktion
            if up_clicked:
                st.session_state.indices[i] = clamp_mod(st.session_state.indices[i] - 1, len(options))
                value_ph.markdown(value_html(options[st.session_state.indices[i]]), unsafe_allow_html=True)

            if down_clicked:
                st.session_state.indices[i] = clamp_mod(st.session_state.indices[i] + 1, len(options))
                value_ph.markdown(value_html(options[st.session_state.indices[i]]), unsafe_allow_html=True)

            value_placeholders.append(value_ph)
    return value_placeholders

def current_values():
    return [CATEGORIES[i][1][st.session_state.indices[i]] for i in range(len(CATEGORIES))]

def render_result(result_ph=None):
    level, leistung, disziplin = current_values()

    unavailable = is_unavailable(level, leistung, disziplin)
    if unavailable:
        title = "nicht verfügbar"
        bg = "#DC2626"  # rot
    else:
        title = compose_title(level, leistung, disziplin)
        bg = THEME["result_bg"]

    html = f'<div class="result-bar" style="background:{bg}">{title}</div>'
    if result_ph is None:
        st.markdown(html, unsafe_allow_html=True)
    else:
        result_ph.markdown(html, unsafe_allow_html=True)

def spin_animation(value_placeholders, result_placeholder, avoid_unavailable=False):
    """Dreht alle Räder nacheinander (Slot-Feeling) mit sanfter Verzögerung und Live-Updates."""
    for i, (_, options) in enumerate(CATEGORIES):
        n = len(options)
        if n == 0:
            continue
        current_idx = st.session_state.indices[i]
        target_idx = random.randrange(n)
        loops = 1 + i
        steps = loops * n + ((target_idx - current_idx) % n)
        if steps == 0:
            steps = n

        base_s = 0.04
        extra_s = 0.09

        for k in range(steps):
            st.session_state.indices[i] = clamp_mod(st.session_state.indices[i] + 1, n)
            val = options[st.session_state.indices[i]]
            value_placeholders[i].markdown(value_html(val), unsafe_allow_html=True)
            render_result(result_placeholder)
            frac = 0 if steps == 1 else k / (steps - 1)
            time.sleep(base_s + frac * extra_s)

        # Optional: verbotene Endkombi überspringen
        if avoid_unavailable:
            max_extra = n * 2
            extra = 0
            while extra < max_extra:
                level, leistung, disziplin = current_values()
                if not is_unavailable(level, leistung, disziplin):
                    break
                st.session_state.indices[i] = clamp_mod(st.session_state.indices[i] + 1, n)
                val = options[st.session_state.indices[i]]
                value_placeholders[i].markdown(value_html(val), unsafe_allow_html=True)
                render_result(result_placeholder)
                time.sleep(0.05)
                extra += 1

# -------------------------
# Sidebar: Regeln laden
# -------------------------
with st.sidebar:
    st.header("Regeln laden")
    uploaded = st.file_uploader("Excel hochladen (nh_kombi_kat2_kat3.xlsx)", type=["xlsx", "xlsm"])
    col_a, col_b = st.columns([1,1])
    with col_a:
        avoid_unavailable = st.checkbox("Spin vermeidet 'nicht verfügbar'", value=False)
    with col_b:
        st.write("")  # spacing
    if uploaded is not None:
        try:
            rules = parse_rules_from_excel(uploaded.read())
            st.session_state.UNAVAILABLE_RULES = rules
            st.success(f"Regeln geladen: {len(rules)} Kombinationen 'nicht verfügbar'.")
        except Exception as e:
            st.error(f"Fehler beim Einlesen der Excel: {e}")

# -------------------------
# Layout
# -------------------------
# Überschrift
st.markdown(
    f"""
    <div style="background:{THEME['bg']};padding:8px 0 0 0;text-align:center;">
      <h1 style="margin:0;color:{THEME['text']};font-weight:800;letter-spacing:.2px;">
        Positionsbezeichnungen bei neues handeln
      </h1>
    </div>
    """,
    unsafe_allow_html=True
)

# Räder
value_placeholders = render_wheels()

# Abstand vor Ergebniszeile
st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)

# Ergebniszeile (zentriert)
result_placeholder = st.empty()
render_result(result_placeholder)

# Mehr Abstand vor den Action-Buttons
st.markdown("<div style='height: 12px;'></div>", unsafe_allow_html=True)

# Action-Buttons zentriert (mit Icons)
outer = st.columns([1, 2, 1], gap="large")
with outer[1]:
    btn_cols = st.columns([1, 1], gap="large")
    with btn_cols[0]:
        st.markdown('<div class="action-btn">', unsafe_allow_html=True)
        spin = st.button("🎰 Spin", type="primary", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with btn_cols[1]:
        st.markdown('<div class="action-btn">', unsafe_allow_html=True)
        reset = st.button("♻️ Reset", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

# Logik
if reset:
    st.session_state.indices = [0 for _ in CATEGORIES]
    value_placeholders = render_wheels()
    render_result(result_placeholder)

if spin:
    spin_animation(value_placeholders, result_placeholder, avoid_unavailable=avoid_unavailable)

